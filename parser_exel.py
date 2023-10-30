import openpyxl as ex
import re
import os
#Путь в папку с нужными нам таблицами в формате эксель, проходим циклом, парсим все нужные данные
#Предварительно создаем эксель фаил пустой, записываем его путь, добавляем список заголовков,
#добавляем все данные в папке
folder_path = '/Users/chicot1k/Desktop/газпром ЦПС'
excel_extensions = ['.xlsx']
processed_files = []
stroka = ''
for filename in os.listdir(folder_path):
    if filename.endswith(tuple(excel_extensions)) and not filename.startswith('.'):
        file_path = os.path.join(folder_path, filename)
        processed_files.append(filename)
        workbook = ex.load_workbook(file_path)
        sheet = workbook.active
        try:
          numb_smeta = sheet['F5'].value
          pattern = r'\d-\d\d-\d{1,2}Р ИНВ\.\d{8}'
          if re.match(pattern, numb_smeta):
              print(f"Номер : {numb_smeta}")
        except:
            found_first_smeta = False
            for row in range(1, sheet.max_row + 1):
                if found_first_smeta:
                    break  # Если нашли первую смету, завершаем цикл
                for column in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=row, column=column).value
                    if isinstance(cell_value, str) and cell_value.startswith("№"):
                        numb_smeta = sheet.cell(row=row, column=column + 1).value
                        print(f"Номер : {numb_smeta}")
                        found_first_smeta = True

        found_work = False
        for row in range(1, sheet.max_row + 1):
            if found_work:
                break  # Если нашли первую смету, завершаем цикл
            for column in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row, column=column).value
                if isinstance(cell_value, str) and cell_value.startswith("(наименование работ "):
                    type_work = sheet.cell(row=row - 1, column=column).value
                    print(f"Наименование {type_work}")
                    found_work = True

        found_work = False
        for row in range(1, sheet.max_row + 1):
            if found_work:
                break  # Если нашли первую смету, завершаем цикл
            for column in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row, column=column).value
                if isinstance(cell_value, str) and cell_value.startswith("Основание:"):
                    reason = sheet.cell(row=row, column=column + 2).value
                    ius = sheet.cell(row=row + 1, column=column + 2).value
                    print(f"Основание: {reason}")
                    print(f"Код : {ius}")
                    found_work = True

        total_row = None
        total_values = None
        flag_value = False
        for row in range(sheet.max_row, 0, -1):
            for column in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row, column=column).value
                if isinstance(cell_value, str) and cell_value.startswith("Итого :"):
                    total_row = row
                    break

            if total_row is not None:
                break


        if total_row is not None:
            total_values = [sheet.cell(row=total_row, column=column).value for column in range(1, sheet.max_column + 1)]
            for total_smeta in total_values:
              if isinstance(total_smeta, int):
                print(f"Итого по смете: {total_smeta}")
                flag_value = True
                break
            if flag_value == False:
              print('Проверьте данные в ручную')
        else:
            print("Итого по смете - не найдена в таблице.")
            total_smeta = None



        total_row = None
        total_values = None
        flag_value = False

        # Начнем поиск с конца таблицы, перебирая строки в обратном порядке
        for row in range(sheet.max_row, 0, -1):
            for column in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row, column=column).value
                if isinstance(cell_value, str) and cell_value.startswith("Нормативная трудоемкость -"):
                    total_row = row
                    break

            if total_row is not None:
                break


        if total_row is not None:
            total_values = [sheet.cell(row=total_row, column=column).value for column in range(1, sheet.max_column + 1)]
            for difficult in total_values:
              if isinstance(difficult, int):
                print(f"Нормативная трудоемкость - {difficult}")
                flag_value = True
                break
            if flag_value == False:
              print('Проверьте данные в ручную')
        else:
            print("Нормативная трудоемкость - не найдена в таблице.")
            difficult = None

        total_row = None
        total_values = None
        flag_value = False


        for row in range(sheet.max_row, 0, -1):
            for column in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row, column=column).value
                if isinstance(cell_value, str) and cell_value.startswith("Всего, стоимость оборудования -"):
                    total_row = row
                    break

            if total_row is not None:
                break  # Нашли строку "Итого по смете:", завершаем поиск


        if total_row is not None:
            total_values = [sheet.cell(row=total_row, column=column).value for column in range(1, sheet.max_column + 1)]
            for total_price_equipment in total_values:
              if isinstance(total_price_equipment, int):
                print(f"Всего, стоимость оборудования - {total_price_equipment}")
                flag_value = True
                break
            if flag_value == False:
              print('Проверьте данные в ручную')
        else:
            print("Всего, стоимость оборудования - не найдена в таблице.")
            total_price_equipment = None

        found_work = False
        amount = None
        metrics = None
        total = None
        for row in range(1, sheet.max_row + 1):
            if found_work:
                break  # Если нашли первую смету, завершаем цикл
            for column in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row, column=column).value
                if isinstance(cell_value, str) and cell_value.startswith("Песчаный карьерный грунт [франко-карьер]"):
                    amount = sheet.cell(row=row, column=column + 1).value
                    metrics = sheet.cell(row=row + 1, column=column + 1).value
                    total = sheet.cell(row=row, column=column + 5).value
                    print(f"Песчаный карьерный грунт [франко-карьер]:")
                    print(f'Количество: {amount} {metrics}')
                    print(f'Итого: {total}')
                    found_work = True  # Устанавливаем флаг, что нашли первую смету

        articles = ['ЛОКАЛЬНАЯ СМЕТА №','Наименование работ и затрат, наименование объекта','Основание','Код ИУС:','Итого по смете:','Нормативная трудоемкость','Всего, стоимость оборудования','Работа','Количество','Единицы измерения','Итог работы']
        numb = [numb_smeta, type_work, reason, ius, total_smeta, difficult, total_price_equipment,stroka, amount, metrics, total]
        path = '/Users/chicot1k/Desktop/test.xlsx'
        new_book = ex.load_workbook(path)
        ws = new_book['data']
        if ws.max_row == 1:
            ws.append(articles)
        ws.append(numb)
        new_book.save(path)
        new_book.close()